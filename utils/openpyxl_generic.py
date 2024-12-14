from openpyxl import Workbook, load_workbook 
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from copy import copy

# FILE_PATH = 'Mention the File path here'
FILE_PATH = '/Users/anirudhan/Documents/All_Documents/Documents/openpyxl/output/'

class ExcelWorkbook():
    '''
    This class is related to the Excel Workbook operations
    '''
    def create_file(self, filename):
        '''
        Creates a file in a specified path
        '''
        complete_file_path = FILE_PATH + filename
        print("Created file including path %s", complete_file_path)
        return complete_file_path

    def create_workbook(self):
        '''
        Creates a workbook
        '''
        print('Inside the create workbook function')
        self.workbook = Workbook()
        return self.workbook
    
    def create_worksheet(self, sheetname):
        '''
        Creates a worksheet with the given name
        '''        
        print('Inside the create sheet function')
        self.sheet = self.workbook.create_sheet(sheetname)
        return self.sheet

    def write_worksheet(self, csv_content):
        '''
        write a worksheet with the given content
        '''   
        print('Inside the write content workbook function')
        for line in csv_content:
            self.sheet.append(line)
        print('Content has been written using workbook function')

    def load_workbook(self, sheetname, complete_file_path):
        '''
        Load a worksheet based on the provided sheetname and its path
        '''   
        print('Inside the load workbook function')
        self.workbook = load_workbook(complete_file_path)
        self.sheet = self.workbook[sheetname]
        return self.sheet

    def save_workbook(self, complete_file_path):
        '''
        Save a workbook in the provided location
        '''          
        print('Saving workbook file in the given path - %s', complete_file_path)
        self.workbook.save(complete_file_path)
        return self.workbook

    def set_sheet_name(self, sheetname):
        '''
        Set a given name to the sheet
        '''          
        print('Set sheet name as given title - %s', sheetname)
        self.sheet.title = sheetname

    def remove_default_sheet(self):
        '''
        To remove default sheet exists in the workbook
        '''          
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        print('Removed default sheet in the workbook - %s', self.workbook)

class FormattingUtils(ExcelWorkbook):
    '''
    This class is related to Formatting based Excel Workbook operations
    '''
    def set_row_col(self):
        '''
        Set min, max - row and column function
        '''
        self.min_row = self.sheet.min_row
        self.header_row = self.sheet.min_row + 1 #Assuming header row will usually be the second line

        self.max_row = self.sheet.max_row
        self.last_row = self.sheet.max_row + 1 #Optional - can be removed if not needed

        self.min_col = self.sheet.min_column
        self.max_col = self.sheet.max_column
        print('min row, min column, max row, max column - %s %s %s %s',
        self.min_row, self.min_col, self.max_row, self.max_col)

    def set_formatting_row_col(self, format_request):
        '''
        Set Formatting - row and column function
        if is_row var - True - > Then formatting occurs for row
        else - > formatting occurs for column
        '''
        if format_request['is_row']:
            self.format_min_row=format_request['input_loc']
            self.format_max_row=format_request['input_loc']+1
            self.format_min_col=format_request['l_range']
            self.format_max_col=format_request['h_range']
        else:
            self.format_min_row=format_request['l_range']
            self.format_max_row=format_request['h_range']
            self.format_min_col=format_request['input_loc']
            self.format_max_col=format_request['input_loc']+1

    def fetch_row_number(self, field, st_row=None, col_num=None):
        '''
        Fetch row number for the given field based on the specified column number
            a. If we need to check and fetch from certain 
            rows -> give st_row (start row), end row is default last row
            b. If we found the row, row number has been returned else None
        Note: We can modify the function to give end row as well if needed
        '''
        print('Fetch row number for the given field')
        start_row = self.min_row if not st_row else st_row
        col_num = self.min_col if not col_num else col_num
        for row_num in range(start_row, self.last_row):
            cell_obj = self.sheet.cell(row=row_num, column=col_num)
            if cell_obj.value==field:
                return row_num
        return None

    def fetch_column_number(self, field, st_col=None, row_num=None):
        '''
        Fetch col number for the given field based on the specified row number
            a. If we need to check and fetch from certain 
            cols -> give st_col (start col), end col is default last col
            b. If we found the col, col number has been returned else None
        Note: We can modify the function to give end col as well if needed
        '''
        print('Fetch column number for the given field')
        start_col = self.min_col if not st_col else st_col
        row_num = self.min_row if not row_num else row_num
        for col_num in range(start_col, self.max_col+1):
            cell_obj = self.sheet.cell(row=row_num, column=col_num)
            if cell_obj.value==field:
                return col_num
        return None

    def merge_cells(self, start_row, start_column, end_row, end_column):
        '''
        Merge cells for the given start, end - row & column
        '''
        print('Inside merge cells function with the given params -> '
        'start row - %s, start column - %s, end row - %s, end column - %s', start_row,
         start_column, end_row, end_column)
        
        self.sheet.merge_cells(start_row=start_row, 
        start_column=start_column, 
        end_row=end_row, 
        end_column=end_column)

class CellStyle(FormattingUtils):
    '''
    This class is related to cell styling based formatting operations
    '''
    custom_style_dict = {
        'bold': {},
        'bold_center': {'horizontal':'center'},
        'bold_center_wrap': {'horizontal':'center', 'wrap':True},
        'vertical_bold_center': {'vertical':'center'},
        'vert_horiz_bold_center': {'horizontal':'center', 'vertical':'center'},
        'vert_horiz_bold_center_wrap': {'horizontal':'center', 'vertical':'center', 'wrap':True},
        'bold_right': {'horizontal':'right'},
    }

    def configure_named_style(self, given_style_name, bold_needed, 
    horizon_align, vertical_align, text_wrap):
        '''
        Using the value of the given params mentioned in the 
        function arguments, named style has been configured
        '''
        style_name = NamedStyle(name=given_style_name)
        style_name.font = Font(bold=bold_needed)
        style_name.alignment = Alignment(horizontal=horizon_align, vertical=vertical_align, wrap_text=text_wrap)
        self.workbook.add_named_style(style_name)

    def set_named_styles(self):
        '''
        Set custom named styles using the var custom_style_dict if not already set
        '''
        print('Set custom named styles to be used by the sheet')

        for name, value in self.custom_style_dict.items():
            if name not in self.workbook.named_styles:
                self.configure_named_style(name, True if 'bold' in name else False,
                horizon_align=value.get('horizontal'), 
                vertical_align=value.get('vertical'), 
                text_wrap=value.get('wrap', False))

    def apply_custom_style(self, cell_obj, style_name):
        '''
        Apply the mentioned custom style to the given cell object
        '''
        cell_obj.style = self.workbook.named_styles[self.workbook.named_styles.index(style_name)]

    def set_style(self, format_name=None):
        '''
        Set style using the given format name
        '''
        for row_num in range(self.format_min_row, self.format_max_row):
            for col_num in range(self.format_min_col, self.format_max_col):
                cell_obj = self.sheet.cell(row=row_num, column=col_num)
                self.apply_custom_style(cell_obj, format_name)

    def make_style_format(self, style_request, format_name):
        '''
        Make custom style format using the given style request, format name
        '''
        self.set_formatting_row_col(style_request)
        self.set_style(format_name)

class CellBorder(FormattingUtils):
    '''
    This class is related to cell border based formatting operations
    '''
    def create_border_obj(self, top_bs=None, bottom_bs=None, right_bs=None, left_bs=None):
        '''
        Using the values of function arguments, border object has been created and returned
        '''
        top = Side(border_style=top_bs)
        bottom = Side(border_style=bottom_bs)
        right = Side(border_style=right_bs)
        left = Side(border_style=left_bs)
        border= Border(top=top, bottom=bottom, right=right, left=left)

        return border

    def set_cell_border(self, cell_obj, border_obj=None):
        '''
        Set border object to the given cell object
        '''
        cell_obj.border = border_obj

    def choose_border(self, border_name):
        '''
        Choose border and its value based on the input
        '''
        border_dict = {
            'top_bottom_medium': ('medium', 'medium'),
            'top_bottom_right_medium': ('medium', 'medium', 'medium'),
            'top_bottom_thin': ('thin', 'thin'),
            'bottom_thin': (None, 'thin'),
            'top_thin_bottom_double': ('thin', 'double')
        }
        return border_dict[border_name]

    def apply_border(self, cell_obj, border_name):
        '''
        Apply border to the given cell object
        '''
        border_obj = self.create_border_obj(*self.choose_border(border_name))
        self.set_cell_border(cell_obj, border_obj)

    def set_border(self, format_name=None):
        '''
        Set the cell object for which the border needs to be applied
        '''
        for row_num in range(self.format_min_row, self.format_max_row):
            for col_num in range(self.format_min_col, self.format_max_col):
                cell_obj = self.sheet.cell(row=row_num, column=col_num)
                self.apply_border(cell_obj, format_name)

    def make_border_format(self, style_request, format_name):
        '''
        Make custom border format based on given style request, format name
        '''
        self.set_formatting_row_col(style_request)
        self.set_border(format_name)

class CellColor(FormattingUtils):
    '''
    This class is related to cell color based formatting operations
    '''
    def set_custom_color(self, color_name):
        '''
        Create custom color and return it
        '''
        print('Set custom color to be used for the sheet')
        ColorFill = PatternFill("solid", fgColor=self.color_to_hex(color_name))
        return ColorFill

    def add_title_color(self, color):
        '''
        Add given color to the title
        '''
        print('Add color to the Sheet Title')
        for cell in range(self.min_col, self.max_col+1):
            self.sheet.cell(row=self.min_row, column=cell).fill = color

    def fill_cell_color(self, row, col, color):
        '''
        Fill cells based on the row, col and color given
        '''
        print('Fill given color - %s to the given row  - %s and column - %s', color, row, col)
        self.sheet.cell(row=row, column=col).fill = color

    def color_to_hex(self, color_name):
        '''
        Get hexadecimal code for the given color
        '''
        hex_dict = {
            "red": "FF0000",
            "lightBlue": "80daeb",
            "blue": "0000FF",
            "green": "008000",
            "lightGreen": "90EE90",
            "yellow": "FFFF00",
            "orange": "FFA500",
            "purple": "800080",
            "pink": "FFC0CB",
            "white": "FFFFFF",
            "black": "000000",
            "gray": "808080",
            "lightGray": "D3D3D3"
        }
        return hex_dict[color_name]

    def set_font_color(self, row, col, color_name):
        '''
        Set font color based on the row, col and color given
        '''
        print('Set font color to be used for the cell')
        self.sheet.cell(row=row, column=col).font = Font(b=True, color=self.color_to_hex(color_name))

class CellFormatting(CellStyle, CellBorder, CellColor):
    '''
    This class basically handles everything related to formatting
    '''
    pass

class CellFormula(ExcelWorkbook):
    '''
    This class is related to the Formula operations applied to the Sheet Cell
    '''
    def sum_formula(self, expression):
        '''
        Basically forming the sum formula based on the row, col, expression
        Sum formula example: "=SUM(E2:E5)"
        '''
        expression = ':' if not expression else expression
        print('Inside Sum Formula function')
        self.sheet.cell(column=self.formula_input_col, row=self.formula_input_row).value = '= SUM('+\
            get_column_letter(self.formula_start_col)+str(self.formula_start_row)+\
                expression+get_column_letter(self.formula_end_col)+str(self.formula_end_row)+')'

    def copy_formula(self, expression=None):
        '''
        Basically forming the copy formula based on the row, col, expression
        Sum formula example: "=SheetName!A1"
        '''
        print('Inside Copy Formula function')
        self.sheet.cell(column=self.formula_input_col, row=self.formula_input_row).value = \
            '='+ self.formula_cp_sheet +'!'+ \
                get_column_letter(self.formula_cp_sheet_input_col)+str(self.formula_cp_sheet_input_row)

    def choose_formula(self, formula_name):
        '''
        Choose the formula based on the formula name
        '''
        formula_dict = {
            'sum': self.sum_formula,
            'copy': self.copy_formula,
        }
        return formula_dict[formula_name]

    def set_formula_row_col(self, formula_request):
        '''
        To set the destined formula input, start, end - rows and cols.
        '''
        self.formula_input_row=formula_request['input_row']
        self.formula_input_col=formula_request['input_col']
        self.formula_start_row=formula_request['start_row']
        self.formula_start_col=formula_request['start_col']
        self.formula_end_row=formula_request['end_row']
        self.formula_end_col=formula_request['end_col']
        self.formula_is_row=formula_request['is_row']

    def set_formula_cp_row_col(self, formula_request):
        '''
        To set the source copy formula input, start, end - rows and cols.
        '''
        self.formula_cp_sheet=formula_request['cp_sheet']
        self.formula_cp_sheet_input_row=formula_request['cp_sheet_input_row']
        self.formula_cp_sheet_input_col=formula_request['cp_sheet_input_col']
        self.formula_cp_sheet_start_row=formula_request['cp_sheet_start_row']
        self.formula_cp_sheet_start_col=formula_request['cp_sheet_start_col']
        self.formula_cp_sheet_end_row=formula_request['cp_sheet_end_row']
        self.formula_cp_sheet_end_col=formula_request['cp_sheet_end_col']

    def set_change_input_row_col(self):
        '''
        Decide to change input row/col based on the formula input vars
        '''
        self.change_input_col= True if not self.formula_input_col else False
        self.change_input_row= True if not self.formula_input_row else False

    def calculate_sum_for_range_of_rows_columns(self, formula_name, expression=None):
        '''
        To calculate the sum for range of rows and columns based on formula name and expression
        '''
        print('Calculate Total for the rows & columns')
        if not self.formula_is_row:
            for col_num in range(self.formula_start_col, self.formula_end_col+1):
                in_col = col_num if self.change_input_col else self.formula_input_col
                self.formula_input_col=in_col
                self.formula_start_col=self.formula_end_col=col_num
                if expression !='-' and self.formula_end_row < self.formula_start_row:
                    self.sheet.cell(column=self.formula_input_col, row=self.formula_input_row).value = 0
                else:
                    self.choose_formula(formula_name)(expression)
        else:
            for row_num in range(self.formula_start_row, self.formula_end_row+1):
                in_row = row_num if self.change_input_row else self.formula_input_row
                self.formula_input_row=in_row
                self.formula_start_row=self.formula_end_row=row_num
                if expression !='-' and self.formula_end_col < self.formula_start_col:
                    self.sheet.cell(column=self.formula_input_col, row=self.formula_input_row).value = 0
                else:
                    self.choose_formula(formula_name)(expression)

    def execute_formula(self, formula_name):
        '''
        To execute formula based on the formula name
        '''
        execute_dict={
            'sum': self.calculate_sum_for_range_of_rows_columns,
            'copy': self.cp_range_of_rows_columns
        }
        return execute_dict[formula_name]

    def apply_formula(self, formula_request, formula_name, expression=None):
        '''
        To apply formula based on the formula request, name and expression
        '''
        self.set_formula_row_col(formula_request)
        self.set_change_input_row_col()
        if formula_name=='copy':
            self.set_formula_cp_row_col(formula_request)
        self.execute_formula(formula_name)(formula_name, expression)

    def cp_range_of_rows_columns(self, formula_name, expression=None):
        '''
        To copy range of rows and columns based on the formula name
        '''
        print('Copy Sum Total from other sheet - %s', self.formula_cp_sheet)
        if not self.formula_is_row:
            for cell, colum in zip(range(self.formula_end_col, self.formula_start_col, -1),\
                 range(self.formula_cp_sheet_start_col, self.formula_cp_sheet_end_col, -1)):
                in_col = cell if self.change_input_col else self.formula_input_col
                self.formula_input_col=in_col
                self.formula_cp_sheet_input_col=colum
                self.choose_formula(formula_name)()
        else:
            for cell, row in zip(range(self.formula_end_row, self.formula_start_row, -1),\
                 range(self.formula_cp_sheet_start_row, self.formula_cp_sheet_end_row, -1)):
                in_row = cell if self.change_input_row else self.formula_input_row
                self.formula_input_row=in_row
                self.formula_cp_sheet_input_row=row
                self.choose_formula(formula_name)()            

class WorkbookProperties(CellFormatting, CellFormula):
    '''
    This class contains all the properties of the workbook
    '''
    def add_filter(self):
        '''
        To add filter to the work sheet
        '''
        print('Add filter to all the header rows')
        self.sheet.auto_filter.ref = get_column_letter(self.min_col)+str(self.header_row)+":"+get_column_letter(self.max_col)+str(self.last_row)

class CopyWholeSheet:
    '''
    This class performs the operation of copying whole sheet to another excel/workbook
    Source of the below code is from Stackoverflow answers : https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl
    '''    
    def __init__(self, source_sheet, target_sheet):
        self.source_sheet=source_sheet
        self.target_sheet=target_sheet

    def copy_sheet_attributes(self):
        self.target_sheet.sheet_format = copy(self.source_sheet.sheet_format)
        self.target_sheet.sheet_properties = copy(self.source_sheet.sheet_properties)
        self.target_sheet.merged_cells = copy(self.source_sheet.merged_cells)
        self.target_sheet.page_margins = copy(self.source_sheet.page_margins)
        self.target_sheet.freeze_panes = copy(self.source_sheet.freeze_panes)

        # set row dimensions
        # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
        for rn in range(len(self.source_sheet.row_dimensions)):
            self.target_sheet.row_dimensions[rn] = copy(self.source_sheet.row_dimensions[rn])

        if self.source_sheet.sheet_format.defaultColWidth is None:
            print('Unable to copy default column wide')
        else:
            self.target_sheet.sheet_format.defaultColWidth = copy(self.source_sheet.sheet_format.defaultColWidth)

        # set specific column width and hidden property
        # we cannot copy the entire column_dimensions attribute so we copy selected attributes
        for key, value in self.source_sheet.column_dimensions.items():
            self.target_sheet.column_dimensions[key].min = copy(self.source_sheet.column_dimensions[key].min)
            self.target_sheet.column_dimensions[key].max = copy(self.source_sheet.column_dimensions[key].max)  
            self.target_sheet.column_dimensions[key].width = copy(self.source_sheet.column_dimensions[key].width) 
            self.target_sheet.column_dimensions[key].hidden = copy(self.source_sheet.column_dimensions[key].hidden)

    def copy_cells(self):
        for (row, col), source_cell in self.source_sheet._cells.items():
            target_cell = self.target_sheet.cell(column=col, row=row)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    def copy_sheet(self):
        self.copy_cells()  # copy all the cell values and styles
        self.copy_sheet_attributes()
